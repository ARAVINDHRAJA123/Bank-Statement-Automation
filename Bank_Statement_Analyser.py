"""
Bank Statement Analyser.py
Tuned for HDFC text-based PDF statements.
"""

import re
import pdfplumber
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_PDF   = "Account Statement.pdf"
OUTPUT_XLSX = "Bank_Statement_Report.xlsx"
ANOMALY_Z   = 2.0

# Column x-boundaries discovered from HDFC PDF structure.
# Each tuple: (col_name, x_start, x_end)
# Words whose x0 falls in [x_start, x_end) are assigned to that column.
HDFC_COLS = [
    ("date",       28,  60),
    ("narration",  60, 280),
    ("ref",       280, 360),
    ("value_date",360, 405),
    ("debit",     405, 490),
    ("credit",    490, 560),
    ("balance",   560, 700),
]

CATEGORY_KEYWORDS = {
    "Food & Dining":     ["adyar ananda", "jjj hotel", "jjj cafe", "chai kings",
                          "gonguraa", "tea time", "hot chips", "swiggy", "zomato",
                          "dominos", "pizza", "kfc", "mcdonalds", "restaurant",
                          "cafe", "food", "biryani", "eat", "burger", "starbucks"],
    "Transport":         ["uber", "ola", "rapido", "petrol", "fuel", "irctc",
                          "bus", "metro", "parking", "toll", "flight", "indigo",
                          "airasia", "redbus", "makemytrip"],
    "Shopping":          ["amazon pay groceries", "jjj store", "max retail",
                          "amazon", "flipkart", "myntra", "meesho", "ajio",
                          "zepto", "blinkit", "bigbasket", "nykaa"],
    "Bills & Utilities": ["lazypay", "jio", "airtel", "vi ", "bsnl", "broadband",
                          "recharge", "postpaid", "dth", "fasttag", "electricity"],
    "Health":            ["pharmacy", "medical", "doctor", "hospital", "apollo",
                          "medplus", "1mg", "netmeds", "diagnostic"],
    "Insurance":         ["icici lom", "icici lombard", "lic", "insurance",
                          "bajaj allianz", "hdfc ergo", "star health"],
    "Entertainment":     ["netflix", "prime", "hotstar", "spotify", "bookmyshow",
                          "inox", "pvr"],
    "Finance & EMI":     ["emi", "loan", "mutual fund", "sip", "nach", "ach",
                          "ecs", "neft", "imps", "rtgs", "credit card", "bajaj"],
    "Salary / Income":   ["salary", "sal ", "payroll", "income", "stipend"],
}

# ── Date helpers ──────────────────────────────────────────────────────────────

DATE_FMTS = ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
             "%d %b %Y", "%d %b %y", "%d-%b-%Y", "%d-%b-%y"]

def parse_date(s: str):
    s = str(s).strip()
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def is_date(s: str) -> bool:
    return parse_date(s) is not None

def parse_amount(s) -> float:
    if not s:
        return 0.0
    cleaned = re.sub(r"[^\d.]", "", str(s))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0

# ── Core extraction ───────────────────────────────────────────────────────────

def assign_col(x0: float) -> str | None:
    for name, x_start, x_end in HDFC_COLS:
        if x_start <= x0 < x_end:
            return name
    return None

def extract_transactions(pdf_path: str) -> list[dict]:
    """
    Spatial extraction using known HDFC column x-positions.
    Groups words by row (y coordinate), assigns to columns, reconstructs
    multi-line narrations by accumulating continuation lines.
    """
    all_rows = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            # Find the header row y — or fall back to the table start zone.
            # Page 2+ may not have a visible header (continuation page).
            header_top = None
            for w in words:
                if w["text"] in ("Date", "Narration"):
                    header_top = w["top"]
                    break

            # If no header found, look for the first date-like token below
            # the address block (address ends by ~y=210) and use that as start.
            if header_top is None:
                for w in sorted(words, key=lambda x: x["top"]):
                    if w["top"] > 215 and is_date(w["text"]) and w["x0"] < 50:
                        header_top = w["top"] - 5  # start just above first row
                        break
            if header_top is None:
                continue  # truly no table on this page

            # Group words into lines by quantised y (±3px tolerance)
            lines: dict[int, list] = {}
            for w in words:
                if w["top"] <= header_top:
                    continue
                y_key = round(w["top"] / 8.6) * 8   # HDFC row height ≈ 17.2pt → bucket by half
                lines.setdefault(y_key, []).append(w)

            # Sort lines top-to-bottom, words left-to-right within each line
            pending: dict | None = None

            for y_key in sorted(lines):
                line_words = sorted(lines[y_key], key=lambda w: w["x0"])

                # Build column buckets for this line
                buckets: dict[str, list[str]] = {n: [] for n, *_ in HDFC_COLS}
                for w in line_words:
                    col = assign_col(w["x0"])
                    if col:
                        buckets[col].append(w["text"])

                date_str  = " ".join(buckets["date"])
                narr_str  = " ".join(buckets["narration"])
                ref_str   = " ".join(buckets["ref"])
                vdt_str   = " ".join(buckets["value_date"])
                dbt_str   = " ".join(buckets["debit"])
                crd_str   = " ".join(buckets["credit"])
                bal_str   = " ".join(buckets["balance"])

                # Skip summary / footer lines
                if any(kw in narr_str for kw in ("OpeningBalance", "ClosingBal",
                                                   "STATEMENTSUMMARYGeneratedOn",
                                                   "STATEMENTSUM", "GeneratedOn")):
                    continue
                if any(kw in date_str for kw in ("DrCount", "OpeningBalance")):
                    continue

                if is_date(date_str):
                    # New transaction row — save previous
                    if pending:
                        all_rows.append(pending)
                    pending = {
                        "date":       parse_date(date_str),
                        "narration":  narr_str,
                        "ref_no":     ref_str,
                        "value_date": parse_date(vdt_str) or parse_date(date_str),
                        "debit":      parse_amount(dbt_str),
                        "credit":     parse_amount(crd_str),
                        "balance":    parse_amount(bal_str),
                    }
                elif pending:
                    # Continuation line — append narration, fill missing amounts
                    if narr_str:
                        pending["narration"] += " " + narr_str
                    if not pending["debit"]   and dbt_str: pending["debit"]   = parse_amount(dbt_str)
                    if not pending["credit"]  and crd_str: pending["credit"]  = parse_amount(crd_str)
                    if not pending["balance"] and bal_str: pending["balance"] = parse_amount(bal_str)

            if pending:
                all_rows.append(pending)

    return all_rows

# ── Enrichment ────────────────────────────────────────────────────────────────

def extract_merchant(narration: str) -> str:
    """Pull a clean merchant name from HDFC UPI / POS narration strings."""
    text = narration.strip()

    # UPI pattern: UPI-MERCHANTNAME-...
    m = re.match(r"UPI-([A-Za-z0-9 &']+?)(?:-[A-Z0-9@]{4,}|-\d|$)", text, re.IGNORECASE)
    if m:
        name = m.group(1).strip().title()
        return name[:40]

    # POS pattern
    m = re.search(r"POS\s+([A-Za-z][A-Za-z0-9 &']+)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    # NEFT/IMPS: take part after bank code
    m = re.search(r"(?:NEFT|IMPS|RTGS)[-/ :]+\S+[-/ :]+([A-Za-z][A-Za-z0-9 ]+)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().title()[:40]

    return text[:40]

def assign_category(narration: str, credit: float) -> str:
    text = narration.lower()
    if credit > 0 and any(k in text for k in ["salary", "sal ", "payroll", "income", "stipend"]):
        return "Salary / Income"
    for cat, keywords in CATEGORY_KEYWORDS.items():
        if any(k.lower() in text for k in keywords):
            return cat
    return "Other Income" if credit > 0 else "Other Expense"

def clean_and_enrich(raw: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for r in raw:
        if r.get("date") is None:
            continue
        key = (r["date"], r["narration"][:30], r["debit"], r["credit"])
        if key in seen:
            continue
        seen.add(key)
        r["merchant"] = extract_merchant(r["narration"])
        r["category"] = assign_category(r["narration"], r["credit"])
        out.append(r)
    out.sort(key=lambda x: x["date"])
    return out

# ── Analytics ─────────────────────────────────────────────────────────────────

def monthly_summary(rows):
    months = {}
    for r in rows:
        key = r["date"].strftime("%b %Y")
        months.setdefault(key, {"month": key, "income": 0.0, "expense": 0.0})
        months[key]["income"]  += r["credit"]
        months[key]["expense"] += r["debit"]
    for m in months.values():
        m["net"] = m["income"] - m["expense"]
    return list(months.values())

def category_summary(rows):
    cats = {}
    for r in rows:
        cat = r["category"]
        cats.setdefault(cat, {"category": cat, "spend": 0.0, "txn_count": 0})
        cats[cat]["spend"]     += r["debit"] if r["debit"] else r["credit"]
        cats[cat]["txn_count"] += 1
    return sorted(cats.values(), key=lambda x: x["spend"], reverse=True)

def top_merchants(rows):
    m = {}
    for r in rows:
        if r["debit"] > 0:
            m[r["merchant"]] = m.get(r["merchant"], 0) + r["debit"]
    return [{"merchant": k, "total_spend": v}
            for k, v in sorted(m.items(), key=lambda x: x[1], reverse=True)[:10]]

def detect_anomalies(rows):
    debits = [r["debit"] for r in rows if r["debit"] > 0]
    if len(debits) < 3:
        return []
    mean = sum(debits) / len(debits)
    std  = (sum((x - mean) ** 2 for x in debits) / len(debits)) ** 0.5
    thresh = mean + ANOMALY_Z * std
    return sorted([r for r in rows if r["debit"] > thresh],
                  key=lambda x: x["debit"], reverse=True)

def spending_stats(rows):
    debits  = [r["debit"]  for r in rows if r["debit"]  > 0]
    credits = [r["credit"] for r in rows if r["credit"] > 0]
    return {
        "total_spend":     sum(debits),
        "total_income":    sum(credits),
        "net_cash_flow":   sum(credits) - sum(debits),
        "avg_expense":     sum(debits) / len(debits) if debits else 0,
        "largest_expense": max(debits)  if debits  else 0,
        "largest_credit":  max(credits) if credits else 0,
        "txn_count":       len(rows),
        "expense_count":   len(debits),
        "income_count":    len(credits),
    }

# ── Excel export ──────────────────────────────────────────────────────────────

C_NAVY  = "1A3C5E"
C_WHITE = "FFFFFF"
C_ALT   = "EFF4F9"
C_RED   = "C0392B"
C_GREEN = "1E8449"
C_BLUE  = "2471A3"
C_WARN  = "FDECEA"
C_WARN_T= "922B21"

THIN   = Side(style="thin", color="CCCCCC")
BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INR    = '₹#,##0.00'
DTEFMT = 'DD/MM/YYYY'

def hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=10)
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = BDR
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def dc(ws, row, col, value, fmt=None, bold=False, color=None, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9, bold=bold, color=color or "000000")
    c.border    = BDR
    c.alignment = Alignment(vertical="center")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    elif row % 2 == 0:
        c.fill = PatternFill("solid", fgColor=C_ALT)
    if fmt:
        c.number_format = fmt
    return c

def write_summary(wb, stats):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20

    t = ws.cell(row=1, column=1, value="HDFC Bank Statement — Analysis")
    t.font = Font(name="Arial", bold=True, size=14, color=C_BLUE)
    ws.merge_cells("A1:B1")

    pairs = [
        ("Total Income",         stats["total_income"],    INR,  C_GREEN),
        ("Total Expenses",       stats["total_spend"],     INR,  C_RED),
        ("Net Cash Flow",        stats["net_cash_flow"],   INR,  C_GREEN if stats["net_cash_flow"] >= 0 else C_RED),
        ("Total Transactions",   stats["txn_count"],       None, C_BLUE),
        ("Expense Transactions", stats["expense_count"],   None, C_RED),
        ("Income Transactions",  stats["income_count"],    None, C_GREEN),
        ("Avg Expense",          stats["avg_expense"],     INR,  C_RED),
        ("Largest Expense",      stats["largest_expense"], INR,  C_RED),
        ("Largest Credit",       stats["largest_credit"],  INR,  C_GREEN),
    ]
    for i, (label, val, fmt, color) in enumerate(pairs, 3):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=True)
        vc = ws.cell(row=i, column=2, value=val)
        vc.font = Font(name="Arial", size=10, bold=(label == "Net Cash Flow"), color=color)
        if fmt:
            vc.number_format = fmt

def write_transactions(wb, rows, anomaly_ids):
    ws = wb.create_sheet("Transactions")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    cols = [("Date",12),("Merchant",24),("Narration",46),("Ref No",20),
            ("Value Date",12),("Debit (₹)",14),("Credit (₹)",14),
            ("Balance (₹)",14),("Category",22),("Flag",8)]
    for i, (h, w) in enumerate(cols, 1):
        hdr(ws, 1, i, h, w)

    for idx, r in enumerate(rows, 2):
        is_anom = id(r) in anomaly_ids
        bg = C_WARN if is_anom else None
        dc(ws, idx, 1, r["date"],       DTEFMT, bg=bg)
        dc(ws, idx, 2, r["merchant"],           bg=bg)
        dc(ws, idx, 3, r["narration"],          bg=bg)
        dc(ws, idx, 4, r["ref_no"],             bg=bg)
        dc(ws, idx, 5, r["value_date"], DTEFMT, bg=bg)
        dc(ws, idx, 6, r["debit"]  or None, INR, color=C_RED   if r["debit"]  > 0 else None, bg=bg)
        dc(ws, idx, 7, r["credit"] or None, INR, color=C_GREEN if r["credit"] > 0 else None, bg=bg)
        dc(ws, idx, 8, r["balance"],        INR, bg=bg)
        dc(ws, idx, 9, r["category"],           bg=bg)
        dc(ws, idx,10, "⚠" if is_anom else "",
           bold=is_anom, color=C_WARN_T if is_anom else None, bg=bg)

    ws.auto_filter.ref = "A1:J1"

def write_monthly(wb, monthly):
    ws = wb.create_sheet("Monthly Summary")
    for i, (h, w) in enumerate([("Month",12),("Income (₹)",16),("Expense (₹)",16),("Net (₹)",16)], 1):
        hdr(ws, 1, i, h, w)
    for idx, m in enumerate(monthly, 2):
        dc(ws, idx, 1, m["month"])
        dc(ws, idx, 2, m["income"],  INR, color=C_GREEN)
        dc(ws, idx, 3, m["expense"], INR, color=C_RED)
        dc(ws, idx, 4, m["net"], INR, bold=True, color=C_GREEN if m["net"] >= 0 else C_RED)

    if monthly:
        from openpyxl.chart.label import DataLabelList
        last = len(monthly) + 1
        chart = BarChart()
        chart.type, chart.grouping = "col", "clustered"
        chart.title = "Monthly Income vs Expense"
        chart.y_axis.title  = "Amount (Rs.)"
        chart.x_axis.title  = "Month"
        chart.y_axis.numFmt = "#,##0"
        chart.width, chart.height, chart.style = 26, 14, 10
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        chart.series[0].graphicalProperties.solidFill = "1E8449"
        chart.series[1].graphicalProperties.solidFill = "C0392B"
        for s in chart.series:
            dlbls = DataLabelList()
            dlbls.showVal = True
            dlbls.showLegendKey = False
            dlbls.showCatName   = False
            dlbls.showSerName   = False
            dlbls.showPercent   = False
            s.dLbls = dlbls
        ws.add_chart(chart, "F2")

def write_categories(wb, cats):
    ws = wb.create_sheet("Categories")
    for i, (h, w) in enumerate([("Category",24),("Total (₹)",16),("Transactions",14)], 1):
        hdr(ws, 1, i, h, w)
    for idx, c in enumerate(cats, 2):
        dc(ws, idx, 1, c["category"])
        dc(ws, idx, 2, c["spend"], INR)
        dc(ws, idx, 3, c["txn_count"])

    if cats:
        last = len(cats) + 1
        pie = PieChart()
        pie.title = "Spending by Category"
        pie.width, pie.height, pie.style = 18, 12, 10
        pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(pie, "E2")

def write_merchants(wb, merchants):
    ws = wb.create_sheet("Top Merchants")
    hdr(ws, 1, 1, "Merchant",           28)
    hdr(ws, 1, 2, "Total Spend (Rs.)",  18)
    for idx, m in enumerate(merchants, 2):
        dc(ws, idx, 1, m["merchant"])
        dc(ws, idx, 2, m["total_spend"], INR)

    if merchants:
        from openpyxl.chart.label import DataLabelList
        last = len(merchants) + 1
        bar = BarChart()
        bar.type  = "bar"
        bar.title = "Top Merchants by Spend"
        bar.y_axis.title = "Merchant"
        bar.x_axis.title = "Amount Spent (Rs.)"
        bar.x_axis.numFmt = "#,##0"
        bar.x_axis.majorGridlines = None
        bar.width, bar.height, bar.style = 28, 18, 10
        bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        bar.series[0].graphicalProperties.solidFill = C_BLUE
        dlbls = DataLabelList()
        dlbls.showVal = True
        dlbls.showLegendKey = False
        dlbls.showCatName   = False
        dlbls.showSerName   = False
        dlbls.showPercent   = False
        bar.series[0].dLbls = dlbls
        ws.add_chart(bar, "D2")

def write_anomalies(wb, anomalies, all_rows):
    ws = wb.create_sheet("Anomalies")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22

    # ── Plain-English explanation block ──────────────────────────────────────
    # Compute stats for the explanation
    debits = [r["debit"] for r in all_rows if r["debit"] > 0]
    mean   = sum(debits) / len(debits) if debits else 0
    std    = (sum((x - mean)**2 for x in debits) / len(debits))**0.5 if debits else 0
    thresh = mean + ANOMALY_Z * std

    title_cell = ws.cell(row=1, column=1,
        value="Unusual Transactions — These payments are much larger than normal")
    title_cell.font = Font(name="Arial", bold=True, size=12, color=C_WARN_T)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 20

    explain_lines = [
        f"How this works: We looked at all {len(debits)} expense transactions in your statement.",
        f"Your typical (average) transaction is Rs. {mean:,.0f}.",
        f"Transactions more than Rs. {thresh:,.0f} are flagged below — these stand out as unusually large.",
        f"This helps you spot: accidental double payments, fraud, or big spends you may have forgotten.",
    ]
    for i, line in enumerate(explain_lines, 2):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="Arial", size=10,
                      color="555555" if i > 2 else "222222",
                      italic=(i > 2))
        ws.merge_cells(f"A{i}:E{i}")
        ws.row_dimensions[i].height = 16

    # Blank separator row
    ws.row_dimensions[6].height = 8

    if not anomalies:
        no_flag = ws.cell(row=7, column=1,
            value="Good news — no unusually large transactions found in this statement.")
        no_flag.font = Font(name="Arial", size=11, color=C_GREEN, bold=True)
        ws.merge_cells("A7:E7")
        return

    # ── Header row ───────────────────────────────────────────────────────────
    for i, (h, w) in enumerate([("Date",12),("Who was paid",28),
                                  ("Full narration",44),
                                  ("Amount paid (Rs.)",18),("Category",22)], 1):
        hdr(ws, 7, i, h)

    # ── Data rows ────────────────────────────────────────────────────────────
    for idx, r in enumerate(anomalies, 8):
        how_big = r["debit"] / mean if mean else 0
        dc(ws, idx, 1, r["date"],       DTEFMT, bg=C_WARN)
        dc(ws, idx, 2, r["merchant"],            bg=C_WARN)
        dc(ws, idx, 3, r["narration"],            bg=C_WARN)
        dc(ws, idx, 4, r["debit"], "#,##0.00",
           bold=True, color=C_WARN_T, bg=C_WARN)
        dc(ws, idx, 5, r["category"],             bg=C_WARN)

    # ── Per-row plain-English reason ─────────────────────────────────────────
    reason_row = len(anomalies) + 9
    ws.cell(row=reason_row, column=1,
            value="Why each transaction above was flagged:").font = Font(
                name="Arial", size=10, bold=True, color="444444")
    ws.merge_cells(f"A{reason_row}:E{reason_row}")

    for i, r in enumerate(anomalies, reason_row + 1):
        how_big = r["debit"] / mean if mean else 0
        reason = (
            f"{r['merchant']} — Rs. {r['debit']:,.0f} on {r['date'].strftime('%d %b %Y')}. "
            f"This is {how_big:.1f}x your average spend of Rs. {mean:,.0f}. "
            f"Category: {r['category']}."
        )
        c = ws.cell(row=i, column=1, value=reason)
        c.font  = Font(name="Arial", size=9, color="444444")
        c.fill  = PatternFill("solid", fgColor="FFF8F7")
        ws.merge_cells(f"A{i}:E{i}")
        ws.row_dimensions[i].height = 18

def export_excel(rows, monthly, cats, merchants, anomalies, stats, path):
    wb = Workbook()
    anomaly_ids = {id(a) for a in anomalies}
    write_summary(wb, stats)
    write_transactions(wb, rows, anomaly_ids)
    write_monthly(wb, monthly)
    write_categories(wb, cats)
    write_merchants(wb, merchants)
    write_anomalies(wb, anomalies, rows)
    wb.save(path)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import os
    pdf = INPUT_PDF
    if not os.path.exists(pdf):
        print(f"ERROR: '{pdf}' not found. Place your HDFC PDF here.")
        return

    print("\n[1/4] Extracting transactions from PDF…")
    raw = extract_transactions(pdf)
    print(f"  Raw rows: {len(raw)}")

    print("[2/4] Cleaning & enriching…")
    rows = clean_and_enrich(raw)
    if not rows:
        print("  No transactions found. Is the PDF text-based (not a scanned image)?")
        return
    print(f"  Valid transactions: {len(rows)}")
    for r in rows:
        print(f"    {r['date']}  {r['merchant'][:30]:<30}  "
              f"debit={r['debit']:>10.2f}  credit={r['credit']:>10.2f}  "
              f"bal={r['balance']:>12.2f}  [{r['category']}]")

    print("[3/4] Running analytics…")
    monthly   = monthly_summary(rows)
    cats      = category_summary(rows)
    merchants = top_merchants(rows)
    anomalies = detect_anomalies(rows)
    stats     = spending_stats(rows)
    print(f"  Periods: {len(monthly)} | Categories: {len(cats)} | Anomalies: {len(anomalies)}")

    print(f"[4/4] Writing {OUTPUT_XLSX}…")
    export_excel(rows, monthly, cats, merchants, anomalies, stats, OUTPUT_XLSX)

    print(f"\n✅  Done! Open '{OUTPUT_XLSX}'")
    print(f"    Transactions  : {stats['txn_count']}")
    print(f"    Total spend   : ₹{stats['total_spend']:,.2f}")
    print(f"    Total income  : ₹{stats['total_income']:,.2f}")
    print(f"    Net cash flow : ₹{stats['net_cash_flow']:,.2f}")
    print(f"    Anomalies     : {len(anomalies)} flagged")

if __name__ == "__main__":
    main()
